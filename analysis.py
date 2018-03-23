# -*- coding: utf-8 -*-
import csv
import json
import argparse
import os
from dbManager import dbManager 
from openpyxl import load_workbook
import operator

class companyAnalysis(object):
    def __init__(self, config_file):
        with open(config_file) as file:
            self.config = json.load(file)

        self.taxcode_index=self.config['taxcode_index']
        self.kind=self.config['kind']

        self.dbManager=dbManager(self.config['db_config'])
            
    def check_folder(self, folder_path):
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)

    def parseThaubing(self):  
        '''
        check coverage of fine_corp and group_realtion data
        ''' 
        input_handler=open(self.factory_corp, 'r')
        reader=csv.DictReader(input_handler)
        factory_set=set()
        for row in reader:
            factory_set.add(row['corp_no'])
        input_handler.close()
        print(len(factory_set))

        input_handler=open(self.company_group, 'r')
        reader=csv.DictReader(input_handler)
        group_set=set()
        for row in reader:
            group_set.add(row['taxcode'])
        print(len(group_set))

        common_set=factory_set.intersection(group_set) 
        print(common_set)  
        print(len(common_set))  

    def generate_group_no(self):
        input_handler=open(self.company_group, 'r')
        reader=csv.DictReader(input_handler)
        
        group_list=set()
        for row in reader:
            group_no=row['group']
            group_list.add(group_no)

        output_handler=open(self.group_list, 'w')
        output_handler.write(json.dumps(list(group_list)))
        output_handler.close()

    def generate_taxcode_list(self):
        file_name="groupdata.taxcode.{}.csv".format(self.config['start'])
        file_path=os.path.join(self.company_group, file_name)
        input_handler=open(file_path, 'r')
        reader=csv.DictReader(input_handler)

        taxcode_list={}
        for row in reader:
            company_name=row['sublist.source']
            company_taxcode=row['taxcode.source']

            if company_name not in taxcode_list:
                taxcode_list[company_name]=company_taxcode

            company_name=row['sublist.target']
            company_taxcode=row['taxcode.target']

            if company_name not in taxcode_list:
                taxcode_list[company_name]=company_taxcode
                            
        output_handler=open(self.taxcode_list, 'w')
        output_handler.write(json.dumps(taxcode_list, ensure_ascii=False))
        output_handler.close()
        
    def generate_taxcode_ronny(self):
        file_path=os.path.join(self.config['factory_folder'],'taxcode_index.csv')
        input_handler=open(file_path, 'r')
        reader=csv.DictReader(input_handler)

        file_path=os.path.join(self.config['factory_folder'],'taxcode_index.json')
        output_handler=open(file_path, 'w')

        taxcode_list={}
        for row in reader:
            if row['type']=='公司':
                taxcode_list[row['#id']]=row['name']

        output_handler.write(json.dumps(taxcode_list, ensure_ascii=False))
        output_handler.close()         

    def setup_company_db(self):
        file_name="groupInfo_{}.json".format(self.config['start'])
        file_path=os.path.join(self.config['groupInfo_folder'], file_name)
        input_handler=open(file_path, 'r')
        reader=json.load(input_handler)

        stock_input=open(self.config['stock_map'])
        stock_map=json.load(stock_input)
    
        for group in reader:
            group_no=group['group_no']
            company_list=group['company_list']

            for company in company_list:
                item={"company_name":company[0], "group_info":json.dumps({self.config['start']:group_no})}
                print(item)
                if company[0] in stock_map:
                    item['stock']=stock_map[company[0]]
                self.dbManager.insert_company(item)    

    def generate_factory_fine_list(self):
        '''
        transfrom fine_corp csv to find_corp.json
        '''
        self.check_folder(self.config['factory_folder'])
        file_path=os.path.join(self.config['factory_folder'],'taxcode_index.json')
        input_handler=open(file_path, 'r')
        taxcode_map=json.load(input_handler)
        inv_taxcode={v:k for k, v in taxcode_map.items()}
        
        file_path=os.path.join(self.config['factory_folder'],'fine_corp.csv')
        fine_handler=open(file_path, 'r')
        reader=csv.DictReader(fine_handler)

        factory_fine_list={}
        for row in reader:
            taxcode=row['corp_id']
            if taxcode!='NULL':
                if taxcode in factory_fine_list:
                    factory_fine_list[taxcode]['record'].append(row)
                    factory_fine_list[taxcode]['record_count']+=1
                    factory_fine_list[taxcode]['money_amount']+=int(row['penalty_money'])
                else:
                    factory_fine_list[taxcode]={"record_count":1, "money_amount":int(row['penalty_money']), "record":[row]}
            else:
                name=row['facility_name']  
                if name in inv_taxcode:
                    taxcode=inv_taxcode[name]
                    if taxcode in factory_fine_list:
                        factory_fine_list[taxcode]['record'].append(row)
                        factory_fine_list[taxcode]['record_count']+=1
                        factory_fine_list[taxcode]['money_amount']+=int(row['penalty_money'])
                    else:
                        factory_fine_list[taxcode]={"record_count":1, "money_amount":int(row['penalty_money']), "record":[row]}        
        
        file_path=os.path.join(self.config['factory_folder'],'fine_corp.json')
        output_handler=open(file_path, 'w')
        output_handler.write(json.dumps(factory_fine_list, ensure_ascii=False))
        output_handler.close()              

    def generate_taxdiscount_list(self):
        '''
        transform xlsx to json 
        '''
        file_path=os.path.join(self.config['taxdiscount_folder'], '104discount.xlsx')
        input_handler=load_workbook(file_path)
        sheet_name_list=input_handler.get_sheet_names()

        for sheet in sheet_name_list:
            taxdiscount=set()
            sheet_content=input_handler.get_sheet_by_name(sheet)
            row_count=sheet_content.max_row
            for index in range(2, row_count):
                taxcode=sheet_content.cell(row=index, column=3).value
                taxdiscount.add(taxcode)

            sheet=int(sheet)+1911
            filename='{}_discount.json'.format(sheet)
            output_path=os.path.join(self.config['taxdiscount_folder'], filename)
            output_handler=open(output_path, 'w')
            output_handler.write(json.dumps(list(taxdiscount),ensure_ascii=False))
            output_handler.close()
            
    def generate_group_list(self):
        file_name="groupdata.taxcode.{}.csv".format(self.config['start'])
        file_path=os.path.join(self.company_group, file_name)
        input_handler=open(file_path, 'r')
        reader=csv.DictReader(input_handler)

        #taxcode map which generated from ronny data
        taxcode_input=open(self.taxcode_list, 'r')
        taxcode_map=json.load(taxcode_input)

        #stock map which combine otc, pub, rotc, sii
        stock_input=open(self.config['stock_map'])
        stock_map=json.load(stock_input)
        inv_stock={v:k for k, v in stock_map.items()}

        #fine_corp record not related with year
        file_path=os.path.join(self.config['factory_folder'],'fine_corp.json')
        fine_handler=open(file_path, 'r')
        factory_fine_list=json.load(fine_handler)

        taxdiscount_file="{}_discount.json".format(self.config['start'])
        taxdiscount_path=os.path.join(self.config['taxdiscount_folder'], taxdiscount_file)
        taxdiscount_handler=open(taxdiscount_path, 'r')
        taxdiscount_list=json.load(taxdiscount_handler)

        group_company_list={}
        group_stock={}
        for row in reader:
            group_no=row['group']
            source_company=row['sublist.source']
            target_company=row['sublist.target']
            market=row['market']
            stock=int(row['stock'])
            if len(source_company)==1:
                #print(row)
                continue
            if len(target_company)==1:
                #print(row)
                continue

            if group_no not in group_company_list:
                group_company_list[group_no]=set([source_company])
                group_company_list[group_no].add(target_company)
            else:
                group_company_list[group_no].add(source_company)    
                group_company_list[group_no].add(target_company) 

            if group_no not in group_stock:
                if stock in inv_stock:
                    group_stock[group_no]=set([inv_stock[stock]])
            else:
                if stock in inv_stock:
                    group_stock[group_no].add(inv_stock[stock])   
                            
        for group_no in group_company_list:    
            has_fine_record=False
            has_discount=False
            fine_money_amount=0
            fine_record_count=0
            company_list=list(group_company_list[group_no])
            updated_company_list=[]
            for company_name in company_list:
                taxcode=taxcode_map[company_name]
                item={"name":company_name, "taxcode":taxcode, "fine_record":{}, "taxdiscount":False}
                
                if taxcode!='NA' and len(taxcode) and int(taxcode) in taxdiscount_list:
                    item["taxdiscount"]=True
                    has_discount=True
                if taxcode in factory_fine_list:
                    item['fine_record']=factory_fine_list[taxcode]
                    fine_money_amount+=factory_fine_list[taxcode]['money_amount']
                    fine_record_count+=factory_fine_list[taxcode]['record_count']
                    has_fine_record=True 
                updated_company_list.append(item)

            group_company_list[group_no]={"company_list":updated_company_list, "has_record":has_fine_record, "has_discount":has_discount, "fine_money_amount":fine_money_amount, "fine_record_count":fine_record_count}
        
        group_info_list=[]
        group_name_list={}
        for group_no in group_company_list:

            group_info_item={"group_no":group_no, "company_list":group_company_list[group_no]['company_list'], 
                "has_fine_record":group_company_list[group_no]['has_record'], 
                "has_discount":group_company_list[group_no]['has_discount'], 
                "fine_money_amount":group_company_list[group_no]['fine_money_amount'], "fine_record_count":group_company_list[group_no]['fine_record_count']}
            if group_no in group_stock:
                group_info_item["group_name_list"]=list(group_stock[group_no])
            else:
                group_info_item["group_name_list"]=[]    
            group_info_list.append(group_info_item)   
            group_name_list[group_no]=group_info_item["group_name_list"]

        file_name="groupInfo_{}.json".format(self.config['start'])
        self.check_folder(self.config['groupInfo_folder'])
        file_path=os.path.join(self.config['groupInfo_folder'], file_name)
        output_handler=open(file_path, 'w')
        output_handler.write(json.dumps(group_info_list, ensure_ascii=False))
        output_handler.close()        

        file_name="groupName_{}.json".format(self.config['start'])
        file_path=os.path.join(self.config['groupInfo_folder'], file_name)
        output_handler=open(file_path, 'w')
        output_handler.write(json.dumps(group_name_list, ensure_ascii=False))
        output_handler.close()      

    def count_company_number(self, folder_path):
        company_count=0
        for dirPath, dirNames, fileNames in os.walk(folder_path):        
                for f in fileNames:
                    if '_2016_list.json' in f:
                       file_path=os.path.join(dirPath, f)
                       input_handler=open(file_path)
                       content=json.load(input_handler) 
                       company_count+=content['company_summery']['company_amount']
        print(company_count)              

    def generate_group_company_folder(self, folder_path):
        group_name_year={}
        group_record_year={}
        filename="groupCommon.json"  
        common_name_handler=open(filename, 'r')
        common_name_map=json.load(common_name_handler)

        for dirPath, dirNames, fileNames in os.walk(folder_path):        
                for f in fileNames:
                    if '.csv' in f:
                        sub_f=f[:-4]
                        file_info=sub_f.split('_')
                        group_name=file_info[0]
                        year=file_info[1]
                        file_path=os.path.join(dirPath, f)
                        try:
                            represent_name, represent_list, group_fine_num=self.generate_group_company_list(file_path, group_name, year)
                            
                        except Exception as e:
                            print(f)
                            raise
                        if year not in group_name_year:
                            group_name_year[year]={}
                        if year not in group_record_year:
                            group_record_year[year]={}
                        
                        if group_name not in group_name_year[year]:
                            #print("{}, {}".format(represent_name, represent_list))
                            group_name_year[year][group_name]={"group_no":group_name, "name":represent_name, "name_list":represent_list, "fine_num":group_fine_num}
                            group_record_year[year][group_name]=group_fine_num

        for key in group_name_year:   
            filename="groupName_{}.json".format(key)
            index_filename="groupNameIndex_{}.json".format(key)
            file_path=os.path.join(self.config['company_folder'], filename)
            index_file_path=os.path.join(self.config['company_folder'], index_filename)
            sorted_group=sorted(group_record_year[key].items(), key=operator.itemgetter(1), reverse=True)
            sort_group_name=[]
            sort_group_index={}
            idx=0
            for group_no, fine_record in sorted_group:
                sort_group_name.append(group_name_year[key][group_no])
                sort_group_index[group_no]=idx
                idx+=1
            output_handler=open(file_path, 'w')
            output_handler.write(json.dumps(sort_group_name, ensure_ascii=False))
            output_handler.close()        
            output_handler=open(index_file_path, 'w')
            output_handler.write(json.dumps(sort_group_index, ensure_ascii=False))
            output_handler.close()    

    def generate_group_company_list(self, file_path, group_name, year):
        input_handler=open(file_path, 'r')
        reader=csv.DictReader(input_handler)
        
        #stock map which combine otc, pub, rotc, sii
        stock_input=open(self.config['stock_map'])
        stock_map=json.load(stock_input)
        inv_stock={v:k for k, v in stock_map.items()}
        
        filename="mops_groupname.json"  
        common_name_handler=open(filename, 'r')
        common_name_map=json.load(common_name_handler)

        group_detail_info={}
        group_company_list=[]
        company_set=set()
        edgelist=[]
        nodelist=[]
        group_penalty_amount=0
        group_has_fine=False
        group_fine_num=0
        fine_company_list=[]
        target_source_map={}
        core_company=set()
        core_company_stock=set()
        for row in reader:
            row['source']=row['source'].replace('臺灣','台灣')
            row['target']=row['target'].replace('臺灣','台灣')
            core_company.add(inv_stock[int(row['stock'])])
            core_company_stock.add(row['stock'])
            item={"has_fine":False, "source":row['source'],"target":row['target'], "taxcode_source":row['taxcode.source'], "taxcode_target":row['taxcode.target'], "holder":row['sublist.holder']}    
            group_company_list.append(item) 
            if row['target'] in target_source_map:
                target_source_map[row['target']]['source_list'].append({"name":row['source'], "holder":row['sublist.holder']})
            else:
                target_source_map[row['target']]={"taxcode":row['taxcode.target'],'source_list':[{"name":row['source'], "holder":row['sublist.holder']}]} 
            if row['source'] not in company_set:
                company_set.add(row['source'])
                nodelist.append({"name":row['source'], 'taxcode':row['taxcode.source'], 'no_holder':False, 'is_core':False})
                if row['source'] in stock_map:
                    source_stock=stock_map[row['source']] 
                else:
                    source_stock='NA'
                
                self.dbManager.insert_company({"year":year, "sublist_source":row['source'], "group":row['group'], "stock":source_stock, "taxcode_source":row['taxcode.source']})
                if row['taxcode.source']!='NA':
                    has_fine, penalty_money_source, record_num=self.dbManager.query_fine_record_by_taxcode(row['taxcode.source'])
                    #has_fine=False
                    if has_fine:
                        group_has_fine=True
                        group_penalty_amount += penalty_money_source
                        group_fine_num+=record_num
                        fine_company_list.append({"taxcode":row['taxcode.source'], "name":row['source'], "fine_num":record_num, "penalty_amount":penalty_money_source})
            if row['target'] not in company_set:
                company_set.add(row['target'])
                nodelist.append({"name":row['target'], 'taxcode':row['taxcode.target'], 'no_holder':False, 'is_core':False})
                if row['target'] in stock_map:
                    target_stock=stock_map[row['target']]
                else:
                    target_stock='NA'    
                self.dbManager.insert_company({"year":year, "sublist_source":row['target'], "group":row['group'], "stock":target_stock, "taxcode_source":row['taxcode.target']})
                if row['taxcode.target']!='NA':
                    has_fine, penalty_money_target, record_num=self.dbManager.query_fine_record_by_taxcode(row['taxcode.target'])       
                    #has_fine=False
                    if has_fine:
                        group_has_fine=True
                        group_penalty_amount += penalty_money_target
                        group_fine_num+=record_num
                        fine_company_list.append({"taxcode":row['taxcode.target'], "name":row['target'], "fine_num":record_num, "penalty_amount":penalty_money_target})       
        
        group_representation_name=''
        group_representation_list=[]
        group_representation_stock=group_name[1:]
        if group_representation_stock in common_name_map:
            group_representation_name += common_name_map[group_representation_stock]
            group_representation_list.append(common_name_map[group_representation_stock])
            count=1
            core_company_stock.remove(group_representation_stock)
        else:    
            count=0
        for stock in list(core_company_stock):
            if str(stock) in common_name_map:
                if count < 2:
                    group_representation_name += common_name_map[str(stock)]
                    count+=1
                group_representation_list.append(common_name_map[str(stock)])

        #check if holder of every parent company is 0
        for node in nodelist:
            if node['name'] in core_company:
                node['is_core']=True
            if node['name'] in target_source_map:
                holder=0
                node['source']=target_source_map[node['name']]['source_list']
                for source in target_source_map[node['name']]['source_list']:
                    if source['holder']!='NA':
                        holder+=float(source['holder']) 
                    else:
                        holder+=1    
                if holder==0:
                    node['no_holder']=True
        
        node_idx= {nodelist[i]['name']:i for i in range(len(nodelist))}
        for info in group_company_list:
            target_idx=node_idx[info['target']]   
            source_idx=node_idx[info['source']]
            if info['holder']!='NA':
                owner_holder='%06f'%(float(info['holder'])*100)
            else:
                owner_holder=info['holder']
            item={"source":source_idx, "target":target_idx, "holder":owner_holder}
            edgelist.append(item)

        group_detail_info['company_summery']={"group_no":group_name, "group_name":group_representation_name, "company_amount":len(list(company_set)), "has_fine":group_has_fine, "fine_company_amount":len(fine_company_list), "fine_record_num":group_fine_num, "fine_penalty_amount":group_penalty_amount}
        group_detail_info['company_list']=group_company_list 
        group_detail_info['fine_company_list']=fine_company_list
        group_detail_info['target_list']=target_source_map   
        
        file_name="{}_{}_list.json".format(group_name, year)
        file_path=os.path.join(self.config['company_folder'], group_name, file_name)
        output_handler=open(file_path, 'w')
        output_handler.write(json.dumps(group_detail_info, ensure_ascii=False))
        output_handler.close()      

        reparsed_content={"y":year, "nodes":nodelist, "links":edgelist}
        file_name="{}_{}_graph.json".format(group_name, year)
        file_path=os.path.join(self.config['company_folder'], group_name, file_name)
        output_handler=open(file_path, "w")
        output_handler.write(json.dumps(reparsed_content, ensure_ascii=False))
        output_handler.close()

        return group_representation_name, group_representation_list, group_fine_num
    
    def generate_stock_list(self, group_name, file_path):
        common_groupname_input=open(self.config['common_groupname'])
        common_groupname_map=json.load(common_groupname_input)

        #stock map which combine otc, pub, rotc, sii
        stock_input=open(self.config['stock_map'])
        stock_map=json.load(stock_input)
        inv_stock={v:k for k, v in stock_map.items()}
        
        group_stock_map={"group_no":group_name}
        group_stock_list=set()
        common_groupname_list=set()
        input_handler=open(file_path, 'r')
        reader=csv.DictReader(input_handler)
        for row in reader:
            if int(row['stock']) in inv_stock:
                group_stock_list.add(row['stock'])
            if row['stock'] in common_groupname_map:
                common_groupname_list.add(common_groupname_map[row['stock']])
        
        group_stock_map['group_stock_list']=list(group_stock_list)
        group_stock_map['common_groupname']=list(common_groupname_list)
        
        return group_stock_map
    
    def compare_common_group_name(self, folder_path):
        group_name_detail=[]
        group_common_name={}
        for dirPath, dirNames, fileNames in os.walk(folder_path):        
                for f in fileNames:
                    if '.csv' in f:
                        sub_f=f[:-4]
                        file_info=sub_f.split('_')
                        group_name=file_info[0]
                        file_path=os.path.join(dirPath, f)                        
                        group_represent_stock=self.generate_stock_list(group_name, file_path)
                        group_name_detail.append(group_represent_stock)
                        if len(group_represent_stock['common_groupname'])>1:
                            print("common_group definition isn't the same, {}".format(group_represent_stock))
                        if len(group_represent_stock['common_groupname'])==1:
                            group_common_name[group_represent_stock['group_no']]=group_represent_stock['common_groupname'][0]
        
        filename="groupCommon.json"  
        output_handler=open(filename, 'w')
        output_handler.write(json.dumps(group_common_name, ensure_ascii=False))
        output_handler.close() 

    def genearte_mops_groupname(self, folder_path):
        for dirPath, dirNames, fileNames in os.walk(folder_path):        
                for f in fileNames:
                    if '.csv' in f:
                        mops_group_dict={}
                        filename_info=f.split('_')
                        file_path=os.path.join(dirPath, f)
                        input_handler=open(file_path, 'r')
                        reader=csv.DictReader(input_handler)
                        for row in reader:
                            stock=row['公司代號']
                            companyName=row['公司名稱']
                            mops_group_dict[stock]=companyName

                        filename="mops_groupname.json"
                        output_handler=open(filename, 'w')
                        output_handler.write(json.dumps(mops_group_dict, ensure_ascii=False))
                        output_handler.close()     
                            
if __name__=="__main__":
    arg_parser = argparse.ArgumentParser()
    arg_parser.add_argument('-t', '--task', metavar='parse', type=str, nargs=1, required=True,
        help='Specify a task to do. (parse)')
    arg_parser.add_argument('-c', '--config', metavar='config.json', type=str, nargs=1, required=True,
        help='Specify a config file')

    args = arg_parser.parse_args()

    if args.task!=None:
        task=args.task[0]

    if args.config!=None:
        config_name=args.config[0]  
    else:
        config_name='config.json'      

    project_dir = os.path.dirname(os.path.abspath(__file__))
    config_file = project_dir + '/project_config/'+ config_name  
    
    with open(config_file) as file:
        project_config=json.load(file)

    analysis=companyAnalysis(config_file) 
    if task=='thaubing':   
        analysis.parseThaubing() 
    elif task=='generate_group_no':
        analysis.generate_group_no()
    elif task=='generate_taxcode':
        analysis.generate_taxcode_list()
    elif task=='generate_group_list':
        analysis.generate_group_list() 
    elif task=='setup_company_db':
        analysis.setup_company_db()       
    elif task=='generate_taxcode_ronny':
        analysis.generate_taxcode_ronny()
    elif task=='parse_factory_fine':
        analysis.generate_factory_fine_list()
    elif task=='generate_taxdiscount_list':
        analysis.generate_taxdiscount_list()
    elif task=='generate_group_company_list':
        file_name="G1101_2016.csv"
        file_path=os.path.join(project_config['company_folder'], 'G1101', file_name)
        print(file_path)
        analysis.generate_group_company_list(file_path, 'G1101', 2016)    
    elif task=='generate_group_company_folder':
        analysis.generate_group_company_folder(project_config['company_folder'])
    elif task=='compare_common_group_name':
        analysis.compare_common_group_name(project_config['company_folder'])
    elif task=='generate_mops_groupname':
        analysis.genearte_mops_groupname(project_config['mops_groupfolder'])
    elif task=='count_company_number':
        analysis.count_company_number(project_config['company_folder'])
    else:
        print("no match task")        